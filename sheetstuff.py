import openpyxl as xl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


def getSheet(file):

    # workbook is the file, sheet is just the page in the file
    workbook = xl.load_workbook(file)
    sheet = workbook["Raw"]

    return sheet

# this is a convenience function that returns the what column name a response for a question should go in (ex. "Requirement - Select the core course, foundation, or skill & perspective for your data." returns "Requirement") The purpose is to make the code more readable and to make it easier to maintain, as changing the form and/or questions should genrally necessitate changing this function only and nothing else
def getDataHeader(QuestionDescriptor):

    try:

        header = ""

        if "GEG" in QuestionDescriptor:
            header = "Goal"
        elif "core course, foundation, or skill & perspective for your data" in QuestionDescriptor:
            header = "Requirement"
        elif "Select the course you taught" in QuestionDescriptor:
            header = "Course"
        elif "type of assessment" in QuestionDescriptor:
            header = "Assessment Type"
        elif "acceptable achievement" in QuestionDescriptor:
            header = "Met"
        if "unacceptable achievement" in QuestionDescriptor:
            header = "Not Met"
        elif "preselected as the course" in QuestionDescriptor:
            header = "Course"
        elif "First Name" in QuestionDescriptor:
            header = "First Name"
        elif "Last Name" in QuestionDescriptor:
            header = "Last Name"
        elif "Email" in QuestionDescriptor:
            header = "Email"

    except Exception as e:
        print(bcolors.FAIL + str(e) + " " + str(QuestionDescriptor) + bcolors.ENDC)

    if header == "":
        print(bcolors.WARNING + "No header specified for question: " + QuestionDescriptor + " Ignoring..." + bcolors.ENDC)
        return None
    
    return header

def transform(sheet, headers):

    # data is a list of dictionarys, where each outer list is a row in the spreadsheet, and the dictionary is the relationship between the headers and the data
    data = []

    # rowCount represents the the number of already existing rows in the final, transformed spreadsheet, which is not nessicarily the same as the current row in the raw spreadsheet

    rowCount = 0

    # iterate through each row (past the first two because those are headers). Each row contains all the data for one person.
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):

        # check that there is data in the first cell of the row, which should mean we're not at the end
        if row[0] == None:
            break

        try:
            print(bcolors.OKBLUE + "Processing data for " + row[17] + " " + row[18] + "..." + bcolors.ENDC)

        except Exception as e:
            print(bcolors.FAIL + str(e) + " " + str(row) + bcolors.ENDC)
            

        noneDict = {header: None for header in headers}
        data.append(noneDict)

        # iterate through each cell in the row
        columnCount = 0

        for cell in row:

            columnCount += 1

            # check that the cell is not empty
            if cell in [None, ""]:
                continue

            # if the cell is "other", skip it because there will be a "please describe" cell that contains the data that we actually want.
            if "other" == str(cell).lower():
                continue

            # get the header for the cell with getDataHeader() and the value of the cell in row 2 in cell's column
            label = sheet.cell(row=2, column=columnCount).value
            header = getDataHeader(label)

            # check that we know where the data goes
            if header == None:
                continue
            

            print(bcolors.OKCYAN + header + ": " + str(cell) + bcolors.ENDC)

            # check that there isn't already data in data[row.index][header]. If there is, increment rowCount to start putting the data in the next row. 
            if data[rowCount][header] not in [None, ""]:
                print(bcolors.HEADER + "Incrementing rowCount because there is already the data '" + str(data[rowCount][header]) + "' in data[" + str(rowCount) + "][" + header + "]" + bcolors.ENDC)
                rowCount += 1
                noneDict = {header: None for header in headers}
                data.append(noneDict)
                
                # copy over first name, last name, and email since those don't change or repeat
                data[rowCount]["First Name"] = data[rowCount-1]["First Name"]
                data[rowCount]["Last Name"] = data[rowCount-1]["Last Name"]
                data[rowCount]["Email"] = data[rowCount-1]["Email"]

            data[rowCount][header] = cell

        rowCount += 1

    return data


def prettyify(wb, sheetName):

    sheet = wb[sheetName]

    data = sheet.values

    # resize the columns by making the widths of the columns be relative to the content (shamelessly stolen from stack overflow and modified to fit this project)
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = len(str(cell))
            else:
                column_widths += [len(str(cell))]
        
    for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
        sheet.column_dimensions[get_column_letter(i)].width = min(column_width//1.4 + 5, 40)

    # set all rows, except the first, to have a height of 40
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 40

    # set everything to use Consolas Mono
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         cell.font = Font(name="Consolas Mono") 

    # make the emails actually be links
    for row in sheet.iter_rows():

        for cell in row:

            # Enable text wrapping for all cells
            cell.alignment = Alignment(wrap_text=True)

            # Check if the cell value looks like an email address
            if isinstance(cell.value, str) and "@" in cell.value and "." in cell.value:
                # Create the hyperlink formula
                hyperlink = f'=HYPERLINK("mailto:{cell.value}", "{cell.value}")'
                cell.value = hyperlink
                # Apply a font style to indicate it's a link
                cell.font = Font(color="0563C1", underline="single")

    return wb


