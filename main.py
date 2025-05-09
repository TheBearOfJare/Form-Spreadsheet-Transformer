import openpyxl as xl
from sheetstuff import *
from webui import *
import webbrowser

# handy colors to differentiate print statements in the terminal
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


def conversion(file):
    
    # the raw data has each submission from the same user in one row, and we want to split it into individual rows for each response, even from the same person. Users who submit multiple responses will generate multiple rows in the final spreadsheet.

    # get the sheet
    sheet = getSheet(file)

    # headers is a list of all the headers in the transformed spreadsheet, data is a list of dictionarys, where each outer list is a row in the spreadsheet, and the dictionary is the relationship between the headers and the data
    headers = ["First Name", "Last Name", "Email", "Requirement", "Course", "Goal", "Assessment Type", "Met", "Not Met"]
    data = transform(sheet=sheet, headers=headers)
    
    # write the data to a new sheet in the same file

    # create the new spreadsheet
    transformed = xl.load_workbook(file)

    # if there's already a sheet called "Transformed", delete it
    if "Transformed" in transformed.sheetnames:
        transformed.remove(transformed["Transformed"])

    # create the new sheet
    sheet = transformed.create_sheet(title="Transformed")

    # make the headers
    for i in range(len(headers)):
        sheet.cell(row=1, column=i+1).value = headers[i]

    # add the data
    for i in range(len(data)):
        for j in range(len(headers)):
            sheet.cell(row=i+2, column=j+1).value = data[i][headers[j]]


    # pretty up the spreadsheet

    prettyified = prettyify(transformed, sheetName="Transformed")

    # save the spreadsheet and we're done
    prettyified.save("Transformed.xlsx")

    print(bcolors.OKGREEN + "Done!" + bcolors.ENDC)


def main():

    # start flask and open loacalhost:5000
    webbrowser.open("http://localhost:5000")
    app.run(debug=True)
    



if __name__ == "__main__":

    main()